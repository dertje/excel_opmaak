import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# TODO add pivottable

 ## CURRENT DATASET FINISHED 03759ned


# tarievenbenchmark excel opmaak:
# calibri (body), 11
# headers: bolt, background-color: orange, accent 2, lighter 80%. HEX='FFF2E6'
# title: bolt, 15, text-color: blue-grey, text2. HEX='44546A'

# get csv file 
# get csv filename
# look up the filename in tablesinfos Identifiers
# get dataset Title that belongs to the Identifier
# add Title to top of the file
# get ShortDescription from Identifier
# add below Title
# remove background excel lines from rows with title and ShortDescription
# add dataset below
# make headers bolt and add background color


def convert_csv_to_xlsx(csv_file):
    filename = Path(csv_file).stem
    df = pd.read_csv(csv_file, delimiter=';')
    excel_file = f"{filename}.xlsx"
    df.to_excel(excel_file, index=False)
    print(f"Step 1: Dataframe is written to excel")

    wb = load_workbook(excel_file)
    ws = wb.active

    # Disable sheet gridlines globally
    ws.sheet_view.showGridLines = False

    title, short_description = format_table(ws, filename)
    nb_of_columns = add_title(ws, title, df)
    add_description(ws, short_description, nb_of_columns)
    wb.save(excel_file)

def format_table(worksheet, identifier):
    worksheet.insert_rows(idx=1, amount=3)
    print(f"Step 2: 3 rows are inserted above dataframe.")
    worksheet.insert_cols(idx=1, amount=1)
    print(f"Step 3: 1 column is inserted to the left of the df")
    
    # After inserting, header row will be row 4
    header_row_idx = 4
    header_row = worksheet[header_row_idx]

    # Load title and description
    with open('example_tables_infos.json', 'r', encoding='utf-8') as file:
        data = json.load(file)

    title, short_description = "", ""
    for entry in data.get('entries', []):
        if entry.get('Identifier') == identifier:
            title = entry.get('Title')
            short_description = entry.get('ShortDescription')
            break
    print(f"Step 5: The found title is: {title} and its short description")

    # Table formatting
    fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
    dashed = Side(border_style="dashed", color="000000")
    dashed_border = Border(left=dashed, right=dashed, top=dashed, bottom=dashed)

    # Detect table bounds dynamically
    min_row = header_row_idx
    max_row = worksheet.max_row
    min_col = 2  # Column B, after inserting one column at A
    max_col = worksheet.max_column

    print(f"Step 6: Table is from row {min_row} to {max_row}, columns {min_col} to {max_col}")

    # Apply background color to header
    for cell in worksheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=min_row):
        for c in cell:
            if c.value is not None:
                c.fill = fill

    # Apply dashed borders only to the dataset area
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:  # Only add borders where there is data
                cell.border = dashed_border

    return title, short_description


def add_title(worksheet, title, df):
    last_col_index = len(df.columns) + 1
    last_col_letter = get_column_letter(last_col_index)

    cell_range = f"B1:{last_col_letter}1"
    worksheet.merge_cells(cell_range)

    worksheet["B1"] = title
    worksheet["B1"].alignment = Alignment(wrap_text=True, vertical='center')
    worksheet["B1"].font = Font(color="44546A", bold=True, size=15)

    thick = Side(border_style="thick", color="5B9BD5")
    header_bottom_border = Border(bottom=thick)

    for col_idx in range(1, last_col_index + 1):
        col_letter = get_column_letter(col_idx)
        worksheet[f"{col_letter}1"].border = header_bottom_border
        worksheet.column_dimensions[col_letter].width = 25

    worksheet.row_dimensions[1].height = 40
    worksheet.column_dimensions["A"].width = 1

    return last_col_index

def add_description(worksheet, short_description, nb_of_columns):
    start_row = 3
    start_col = 2
    max_cols = 6
    words_per_line = 10

    words = short_description.split()
    lines = [words[i:i+words_per_line] for i in range(0, len(words), words_per_line)]
    worksheet.insert_rows(idx=3, amount=len(lines))
    nb_of_rows_to_freeze = str(5 + len(lines))
    worksheet.freeze_panes = f"C{nb_of_rows_to_freeze}"

    row = start_row
    for line in lines:
        line_text = " ".join(line)
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(start_col + max_cols - 1)
        merge_range = f"{start_letter}{row}:{end_letter}{row}"

        worksheet.merge_cells(merge_range)
        cell = worksheet.cell(row=row, column=start_col)
        cell.value = line_text
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        row += 1


def main():
    convert_csv_to_xlsx('test2.csv')


if __name__ == '__main__':
    main()
